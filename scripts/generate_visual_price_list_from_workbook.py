#!/usr/bin/env python3
"""
Generate polished customer-facing price lists from the Arabic price workbook.

Outputs separate files:
- General USD list
- General SYP list
- Waziri USD list
- Waziri SYP list

Optional inventory filtering:
- If --inventory-file is provided, only items that exist in the inventory export
  and have quantity greater than --min-quantity are published.
- A mismatch report is generated to show:
  - items in price list but not in inventory
  - items in inventory but not in price list

Design rules:
- Black and gold branding.
- A4 landscape.
- Balanced multi-column layout to avoid one very long list.
- No group/code/quantity/packaging columns in the published list.
- Packaging/conversion factor is used internally only.

The original Excel files are read-only and never modified.
"""

from __future__ import annotations

import argparse
import csv
import html
import re
import shutil
import subprocess
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, DefaultDict, Iterable, List, Literal, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: python -m pip install -r requirements.txt"
    ) from exc

from generate_price_list_from_price_workbook import (
    DEFAULT_SHEET_NAME,
    PriceItem,
    load_items_from_workbook,
    parse_date,
    syp_money,
)


BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_DIR = BASE_DIR / "output"
BRAND_NAME = "مركز أبو زياد لتجارة التبغ والدخان الوطني والأجنبي والمستورد"
PHONE_LINES = ["0985000771", "0984000662", "0994092038"]

CurrencyMode = Literal["usd", "syp"]


@dataclass
class InventoryFilterResult:
    published_items: List[PriceItem]
    price_items_not_in_inventory: List[str]
    inventory_items_not_in_price_list: List[str]
    inventory_count: int
    inventory_source: str


def money(value: float) -> str:
    return f"{value:,.2f}"


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def parse_number(value: Any) -> Optional[float]:
    raw = text(value).replace(",", "").strip()
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def normalize_key(value: str) -> str:
    value = value.strip().lower()
    value = value.replace("ـ", "")
    value = re.sub(r"[\s\-_/\\.,،:;؛()\[\]{}]+", "", value)
    value = value.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    value = value.replace("ة", "ه")
    return value


def find_browser_executable() -> str | None:
    candidates = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return candidate
    for name in ["msedge", "chrome", "chrome.exe", "msedge.exe"]:
        found = shutil.which(name)
        if found:
            return found
    return None


def is_waziri(item: PriceItem, keyword: str) -> bool:
    keyword = keyword.strip()
    if not keyword:
        return False
    return keyword in item.group or keyword in item.name


def item_unit_and_price(item: PriceItem, exchange_rate: float, mode: CurrencyMode) -> tuple[str, str]:
    if item.factor > 1:
        if mode == "usd":
            return "كرتونة", money(item.price_usd or 0)
        return "كروز", syp_money((item.price_usd or 0) / item.factor * exchange_rate)
    if mode == "usd":
        return "قطعة", money(item.price_usd or 0)
    return "قطعة", syp_money((item.price_usd or 0) * exchange_rate)


def group_items(items: Iterable[PriceItem]) -> dict[str, List[PriceItem]]:
    grouped: DefaultDict[str, List[PriceItem]] = defaultdict(list)
    for item in items:
        grouped[item.group or "غير مصنف"].append(item)
    return dict(sorted(grouped.items(), key=lambda pair: pair[0]))


def build_cards(items: List[PriceItem], exchange_rate: float, mode: CurrencyMode) -> str:
    grouped = group_items(items)
    if not grouped:
        return '<div class="empty">لا توجد أصناف قابلة للنشر</div>'

    cards: List[str] = []
    for group_name, group_items_list in grouped.items():
        rows: List[str] = []
        for item in sorted(group_items_list, key=lambda item: item.name):
            unit, price = item_unit_and_price(item, exchange_rate, mode)
            rows.append(
                '<div class="item-row">'
                f'<span class="item-name">{html.escape(item.name)}</span>'
                f'<span class="item-unit">{html.escape(unit)}</span>'
                f'<span class="item-price">{html.escape(price)}</span>'
                '</div>'
            )
        cards.append(
            '<section class="group-card">'
            f'<h2>{html.escape(group_name)}</h2>'
            '<div class="rows">'
            + "\n".join(rows)
            + '</div></section>'
        )
    return "\n".join(cards)


def build_review_html(review_items: List[str]) -> str:
    if not review_items:
        return '<div class="review-ok">لا توجد أصناف تحتاج تسعير.</div>'
    items = "\n".join(f"<li>{html.escape(item)}</li>" for item in review_items)
    return f"<ul>{items}</ul>"


def load_inventory_from_csv(file_path: Path, min_quantity: float) -> dict[str, str]:
    with file_path.open("r", encoding="utf-8-sig", newline="") as file:
        rows = list(csv.DictReader(file))
    if not rows:
        return {}

    headers = list(rows[0].keys())
    name_header = find_name_header(headers)
    qty_header = find_quantity_header(headers)
    if not name_header:
        raise ValueError("لم يتم العثور على عمود اسم الصنف داخل ملف المخزون CSV.")

    inventory: dict[str, str] = {}
    for row in rows:
        name = text(row.get(name_header, ""))
        if not name:
            continue
        quantity = parse_number(row.get(qty_header, "")) if qty_header else 1
        if quantity is None:
            quantity = 1
        if quantity > min_quantity:
            inventory[normalize_key(name)] = name
    return inventory


def find_name_header(headers: Iterable[str]) -> Optional[str]:
    candidates = ["اسمالماده", "اسمالصنف", "الصنف", "الماده", "الاسم", "اسم"]
    for header in headers:
        normalized = normalize_key(header)
        if any(candidate in normalized for candidate in candidates):
            return header
    return None


def find_quantity_header(headers: Iterable[str]) -> Optional[str]:
    candidates = ["الكميه", "المتوفر", "الرصيد", "المخزون", "باقي", "الكمية"]
    for header in headers:
        normalized = normalize_key(header)
        if any(candidate in normalized for candidate in candidates):
            return header
    return None


def get_sheet_bounds(sheet: Any) -> tuple[int, int]:
    """Return safe max_row/max_column values, handling empty/read-only worksheets."""
    max_row = sheet.max_row or 0
    max_column = sheet.max_column or 0
    return int(max_row), int(max_column)


def load_inventory_from_excel(file_path: Path, sheet_name: Optional[str], min_quantity: float) -> dict[str, str]:
    # read_only=False is more reliable for some accounting exports whose dimensions are not set.
    workbook = load_workbook(filename=file_path, read_only=False, data_only=True)
    sheet_names = [sheet_name] if sheet_name else workbook.sheetnames
    inventory: dict[str, str] = {}

    for current_sheet_name in sheet_names:
        if not current_sheet_name or current_sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[current_sheet_name]
        max_row, max_column = get_sheet_bounds(sheet)
        if max_row <= 0 or max_column <= 0:
            continue

        max_scan_rows = min(max_row, 80)
        header_row_index: Optional[int] = None
        name_col_index: Optional[int] = None
        qty_col_index: Optional[int] = None

        for row_idx in range(1, max_scan_rows + 1):
            headers = [text(sheet.cell(row_idx, col_idx).value) for col_idx in range(1, max_column + 1)]
            name_header = find_name_header(headers)
            if not name_header:
                continue
            header_row_index = row_idx
            name_col_index = headers.index(name_header) + 1
            qty_header = find_quantity_header(headers)
            qty_col_index = headers.index(qty_header) + 1 if qty_header else None
            break

        if header_row_index is None or name_col_index is None:
            continue

        for row_idx in range(header_row_index + 1, max_row + 1):
            name = text(sheet.cell(row_idx, name_col_index).value)
            if not name:
                continue
            quantity = 1.0
            if qty_col_index is not None:
                parsed_quantity = parse_number(sheet.cell(row_idx, qty_col_index).value)
                quantity = parsed_quantity if parsed_quantity is not None else 0.0
            if quantity > min_quantity:
                inventory[normalize_key(name)] = name

    workbook.close()
    if not inventory:
        raise ValueError(
            "لم يتم العثور على أصناف مخزون صالحة داخل ملف الجرد. "
            "تأكد أن الملف يحتوي على أعمدة: اسم المادة / الكمية."
        )
    return inventory


def load_inventory_names(file_path: Path, sheet_name: Optional[str], min_quantity: float) -> dict[str, str]:
    if not file_path.exists():
        raise FileNotFoundError(f"Inventory file not found: {file_path}")
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return load_inventory_from_csv(file_path, min_quantity)
    if suffix in {".xlsx", ".xlsm"}:
        return load_inventory_from_excel(file_path, sheet_name, min_quantity)
    raise ValueError("Inventory file must be .xlsx, .xlsm, or .csv")


def filter_items_by_inventory(
    items: List[PriceItem],
    inventory_file: Optional[Path],
    inventory_sheet: Optional[str],
    min_quantity: float,
) -> InventoryFilterResult:
    if inventory_file is None:
        return InventoryFilterResult(
            published_items=items,
            price_items_not_in_inventory=[],
            inventory_items_not_in_price_list=[],
            inventory_count=0,
            inventory_source="لم يتم تمرير ملف مخزون؛ تم استخدام ملف الأسعار فقط.",
        )

    inventory_names = load_inventory_names(inventory_file, inventory_sheet, min_quantity)
    price_keys = {normalize_key(item.name): item.name for item in items}
    published_items = [item for item in items if normalize_key(item.name) in inventory_names]
    price_items_not_in_inventory = [item.name for item in items if normalize_key(item.name) not in inventory_names]
    inventory_items_not_in_price_list = [
        display_name for key, display_name in inventory_names.items() if key not in price_keys
    ]

    return InventoryFilterResult(
        published_items=published_items,
        price_items_not_in_inventory=sorted(price_items_not_in_inventory),
        inventory_items_not_in_price_list=sorted(inventory_items_not_in_price_list),
        inventory_count=len(inventory_names),
        inventory_source=str(inventory_file),
    )


def build_html(
    items: List[PriceItem],
    review_items: List[str],
    report_date: date,
    exchange_rate: float,
    source_file: Path,
    title: str,
    mode: CurrencyMode,
    include_review: bool,
) -> str:
    cards = build_cards(items, exchange_rate, mode)
    review_html = build_review_html(review_items)
    generated_at = datetime.now().isoformat(timespec="seconds")
    phones = " | ".join(PHONE_LINES)
    price_label = "السعر بالدولار" if mode == "usd" else "السعر بالليرة السورية"

    return f"""<!doctype html>
<html lang="ar" dir="rtl">
<head>
<meta charset="utf-8">
<title>{html.escape(title)}</title>
<style>
  @page {{ size: A4 landscape; margin: 7mm; }}
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; background: #111; color: #141414; font-family: "Segoe UI", Tahoma, Arial, sans-serif; direction: rtl; }}
  .page {{ width: 297mm; min-height: 210mm; margin: 0 auto; background: #fbfaf7; padding: 8mm; }}
  .header {{ border: 2px solid #c9a227; border-radius: 16px; background: #080808; color: #f7e7a3; padding: 10px 16px; margin-bottom: 8px; }}
  .brand {{ font-size: 23px; font-weight: 900; margin: 0 0 4px 0; line-height: 1.25; }}
  .subline {{ display: flex; justify-content: space-between; gap: 12px; flex-wrap: wrap; color: #fff7d1; font-size: 12px; }}
  .title-bar {{ display: flex; align-items: center; justify-content: space-between; gap: 10px; background: #c9a227; color: #080808; border-radius: 12px; padding: 7px 12px; margin-bottom: 8px; font-weight: 900; }}
  .title-bar h1 {{ font-size: 18px; margin: 0; }}
  .title-bar .meta {{ font-size: 12px; white-space: nowrap; }}
  .legend {{ display: grid; grid-template-columns: minmax(0, 1fr) 58px 90px; gap: 6px; padding: 5px 8px; margin-bottom: 5px; border-radius: 8px; background: #1f1f1f; color: #f7e7a3; font-weight: 900; font-size: 10.5px; }}
  .cards {{ columns: 3; column-gap: 8px; }}
  .group-card {{ break-inside: avoid; page-break-inside: avoid; display: inline-block; width: 100%; border: 1px solid #d8c277; border-radius: 10px; overflow: hidden; margin: 0 0 8px 0; background: #fff; }}
  .group-card h2 {{ margin: 0; padding: 5px 8px; background: #080808; color: #f7e7a3; font-size: 12px; font-weight: 900; }}
  .item-row {{ display: grid; grid-template-columns: minmax(0, 1fr) 58px 90px; gap: 6px; padding: 4px 8px; border-bottom: 1px solid #eee4bd; font-size: 10.2px; line-height: 1.2; }}
  .item-row:last-child {{ border-bottom: 0; }}
  .item-name {{ font-weight: 700; overflow-wrap: anywhere; }}
  .item-unit {{ color: #444; white-space: nowrap; }}
  .item-price {{ direction: ltr; text-align: left; font-weight: 900; white-space: nowrap; }}
  .empty {{ border: 1px solid #d8c277; border-radius: 10px; padding: 16px; background: #fff; font-weight: 900; }}
  .review {{ margin-top: 8px; border: 1px dashed #c9a227; border-radius: 10px; padding: 8px 10px; background: #fff8d8; font-size: 10.5px; break-inside: avoid; }}
  .review h2 {{ margin: 0 0 5px 0; font-size: 12px; }}
  .review ul {{ margin: 0; padding-right: 18px; columns: 3; }}
  .review-ok {{ color: #166534; font-weight: 900; }}
  .footer {{ margin-top: 6px; text-align: center; color: #4b5563; font-size: 9.5px; }}
  .internal {{ display: none; }}
  @media screen and (max-width: 900px) {{ body {{ background: #fbfaf7; }} .page {{ width: auto; min-height: auto; padding: 10px; }} .cards {{ columns: 1; }} .subline, .title-bar {{ display: block; }} }}
  @media print {{ body {{ background: #fff; }} .page {{ width: auto; min-height: auto; padding: 0; }} }}
</style>
</head>
<body>
<div class="page">
  <header class="header">
    <h1 class="brand">{html.escape(BRAND_NAME)}</h1>
    <div class="subline"><span>هاتف: {html.escape(phones)}</span><span>نشرة أسعار منظمة للزبائن</span></div>
  </header>
  <div class="title-bar"><h1>{html.escape(title)}</h1><div class="meta">التاريخ: {html.escape(report_date.isoformat())} — سعر الصرف: {html.escape(syp_money(exchange_rate))}</div></div>
  <div class="legend"><span>الصنف</span><span>الوحدة</span><span>{html.escape(price_label)}</span></div>
  <main class="cards">{cards}</main>
  {f'<section class="review"><h2>مراجعة داخلية قبل النشر</h2>{review_html}</section>' if include_review else ''}
  <div class="footer">الأسعار قابلة للتغيير حسب التوفر وسعر الصرف. يرجى التأكد قبل اعتماد الطلب.</div>
  <div class="internal">Source: {html.escape(str(source_file))} | Generated: {html.escape(generated_at)}</div>
</div>
</body>
</html>
"""


def write_pdf_with_browser(html_path: Path, pdf_path: Path) -> bool:
    browser = find_browser_executable()
    if not browser:
        return False
    file_url = html_path.resolve().as_uri()
    command = [browser, "--headless", "--disable-gpu", f"--print-to-pdf={pdf_path}", file_url]
    subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
    return pdf_path.exists()


def write_one_list(
    items: List[PriceItem],
    review_items: List[str],
    output_dir: Path,
    report_date: date,
    exchange_rate: float,
    source_file: Path,
    title: str,
    mode: CurrencyMode,
    filename_slug: str,
    make_pdf: bool,
    include_review: bool = False,
) -> tuple[Path, Path | None]:
    html_path = output_dir / f"{filename_slug}-{report_date.isoformat()}.html"
    pdf_path = output_dir / f"{filename_slug}-{report_date.isoformat()}.pdf"
    content = build_html(items, review_items, report_date, exchange_rate, source_file, title, mode, include_review)
    html_path.write_text(content, encoding="utf-8")
    created_pdf = None
    if make_pdf and write_pdf_with_browser(html_path, pdf_path):
        created_pdf = pdf_path
    return html_path, created_pdf


def write_inventory_report(output_dir: Path, report_date: date, result: InventoryFilterResult) -> Path:
    path = output_dir / f"inventory-price-mismatch-{report_date.isoformat()}.txt"
    lines = [
        "Inventory and price list mismatch report",
        f"Inventory source: {result.inventory_source}",
        f"Inventory items detected: {result.inventory_count}",
        f"Published items after inventory filter: {len(result.published_items)}",
        "",
        "Items in price workbook but not available in inventory:",
        *(f"- {name}" for name in result.price_items_not_in_inventory[:500]),
        "",
        "Items available in inventory but missing from price workbook:",
        *(f"- {name}" for name in result.inventory_items_not_in_price_list[:500]),
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def generate_visual_price_lists(
    file_path: Path,
    output_dir: Path,
    sheet_name: str,
    report_date: date,
    exchange_rate: float,
    make_pdf: bool,
    waziri_keyword: str,
    inventory_file: Optional[Path],
    inventory_sheet: Optional[str],
    min_quantity: float,
) -> list[Path]:
    items, review_items = load_items_from_workbook(file_path, sheet_name)
    output_dir.mkdir(parents=True, exist_ok=True)

    inventory_result = filter_items_by_inventory(items, inventory_file, inventory_sheet, min_quantity)
    items_for_publication = inventory_result.published_items

    general_items = [item for item in items_for_publication if not is_waziri(item, waziri_keyword)]
    waziri_items = [item for item in items_for_publication if is_waziri(item, waziri_keyword)]

    generated: list[Path] = []
    specs = [
        (general_items, "نشرة الدولار", "usd", "price-list-dollar", True),
        (general_items, "نشرة السوري", "syp", "price-list-syrian", False),
        (waziri_items, "نشرة الوزاري - دولار", "usd", "price-list-waziri-dollar", False),
        (waziri_items, "نشرة الوزاري - سوري", "syp", "price-list-waziri-syrian", False),
    ]
    for list_items, title, mode, slug, include_review in specs:
        html_path, pdf_path = write_one_list(
            items=list_items,
            review_items=review_items,
            output_dir=output_dir,
            report_date=report_date,
            exchange_rate=exchange_rate,
            source_file=file_path,
            title=title,
            mode=mode,  # type: ignore[arg-type]
            filename_slug=slug,
            make_pdf=make_pdf,
            include_review=include_review,
        )
        generated.append(html_path)
        if pdf_path:
            generated.append(pdf_path)

    mismatch_path = write_inventory_report(output_dir, report_date, inventory_result)
    generated.append(mismatch_path)

    summary_path = output_dir / f"price-list-visual-summary-{report_date.isoformat()}.txt"
    summary_path.write_text(
        "\n".join(
            [
                "Visual price list generation summary",
                f"Source price file: {file_path}",
                f"Inventory source: {inventory_result.inventory_source}",
                f"Inventory items detected: {inventory_result.inventory_count}",
                f"Published items after inventory filter: {len(items_for_publication)}",
                f"General items: {len(general_items)}",
                f"Waziri items: {len(waziri_items)}",
                f"Review items without prices: {len(review_items)}",
                f"Price items not in inventory: {len(inventory_result.price_items_not_in_inventory)}",
                f"Inventory items not in price list: {len(inventory_result.inventory_items_not_in_price_list)}",
                f"Exchange rate: {exchange_rate}",
                "Generated files:",
                *[str(path) for path in generated],
                "Original Excel files were not modified.",
            ]
        ),
        encoding="utf-8",
    )
    generated.append(summary_path)
    return generated


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate polished separate HTML/PDF price lists from the price workbook.")
    parser.add_argument("--file", required=True, help="Path to the .xlsx/.xlsm price workbook.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Worksheet name to read.")
    parser.add_argument("--date", default=date.today().isoformat(), help="Report date in YYYY-MM-DD format.")
    parser.add_argument("--exchange-rate", required=True, type=float, help="USD to SYP exchange rate.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Directory for generated output.")
    parser.add_argument("--pdf", action="store_true", help="Also try to create PDFs using Edge/Chrome headless.")
    parser.add_argument("--waziri-keyword", default="وزاري", help="Keyword used to split Waziri items into separate lists.")
    parser.add_argument("--inventory-file", default=None, help="Optional inventory export file from accounting software: .xlsx/.xlsm/.csv")
    parser.add_argument("--inventory-sheet", default=None, help="Optional inventory worksheet name.")
    parser.add_argument("--min-quantity", default=0.0, type=float, help="Minimum quantity required for publishing an item.")
    args = parser.parse_args()

    paths = generate_visual_price_lists(
        file_path=Path(args.file),
        output_dir=Path(args.output_dir),
        sheet_name=args.sheet,
        report_date=parse_date(args.date),
        exchange_rate=args.exchange_rate,
        make_pdf=args.pdf,
        waziri_keyword=args.waziri_keyword,
        inventory_file=Path(args.inventory_file) if args.inventory_file else None,
        inventory_sheet=args.inventory_sheet,
        min_quantity=args.min_quantity,
    )
    print("Visual price lists generated:")
    for path in paths:
        print(f"- {path}")
    print("Original Excel files were not modified.")


if __name__ == "__main__":
    main()
