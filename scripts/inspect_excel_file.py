#!/usr/bin/env python3
"""
Safely inspect an Excel workbook.

This script reads an Excel file in read-only mode and writes a local Markdown
inspection report to output/. It does not modify the Excel file, upload data,
send messages, or connect to external services.

Supported reading engines:
- openpyxl for .xlsx/.xlsm files

Note: legacy .xls files are not supported by the Python standard library or
openpyxl. Convert .xls to .xlsx first using Excel or LibreOffice.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, List, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - user environment guard
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: python -m pip install openpyxl"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_DIR = BASE_DIR / "output"


@dataclass
class SheetInspection:
    sheet_name: str
    max_row: int
    max_column: int
    headers: List[str]
    preview_rows: List[List[str]]


def safe_slug(value: str) -> str:
    value = value.strip()
    value = re.sub(r"[^\w\-.\u0600-\u06FF]+", "-", value, flags=re.UNICODE)
    value = value.strip("-")
    return value or "excel-file"


def cell_to_text(value: Any, max_length: int = 80) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    if len(text) > max_length:
        return text[: max_length - 3] + "..."
    return text


def row_to_texts(row: Iterable[Any]) -> List[str]:
    return [cell_to_text(cell.value) for cell in row]


def markdown_table(headers: List[str], rows: List[List[str]]) -> str:
    if not headers:
        return "لا توجد أعمدة ظاهرة."
    safe_headers = [header or f"Column {index + 1}" for index, header in enumerate(headers)]
    lines = [
        "| " + " | ".join(safe_headers) + " |",
        "| " + " | ".join("---" for _ in safe_headers) + " |",
    ]
    for row in rows:
        padded = row[: len(safe_headers)] + [""] * max(len(safe_headers) - len(row), 0)
        escaped = [value.replace("|", "\\|") for value in padded]
        lines.append("| " + " | ".join(escaped) + " |")
    return "\n".join(lines)


def inspect_sheet(sheet: Any, preview_rows_count: int) -> SheetInspection:
    rows_iter = sheet.iter_rows()
    first_row = next(rows_iter, None)
    headers = row_to_texts(first_row) if first_row else []

    preview_rows: List[List[str]] = []
    for index, row in enumerate(rows_iter):
        if index >= preview_rows_count:
            break
        preview_rows.append(row_to_texts(row))

    return SheetInspection(
        sheet_name=sheet.title,
        max_row=sheet.max_row,
        max_column=sheet.max_column,
        headers=headers,
        preview_rows=preview_rows,
    )


def inspect_workbook(file_path: Path, output_dir: Path, preview_rows: int) -> Path:
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Only .xlsx and .xlsm files are supported. Convert .xls to .xlsx first.")

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    inspections = [inspect_sheet(workbook[sheet_name], preview_rows) for sheet_name in workbook.sheetnames]
    workbook.close()

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"excel-inspection-{safe_slug(file_path.stem)}-{date.today().isoformat()}.md"

    sections: List[str] = [
        "# تقرير فحص ملف Excel",
        "",
        f"- الملف: `{file_path}`",
        f"- وقت الفحص: {datetime.now().isoformat(timespec='seconds')}",
        f"- عدد الأوراق: {len(inspections)}",
        "",
        "## ملاحظة أمان",
        "",
        "- تم فتح الملف للقراءة فقط.",
        "- لم يتم تعديل ملف Excel الأصلي.",
        "- لم يتم إرسال البيانات أو رفعها لأي خدمة خارجية.",
        "- التقرير يعرض عينة محدودة فقط للمراجعة.",
        "",
    ]

    for inspection in inspections:
        sections.extend(
            [
                f"## الورقة: {inspection.sheet_name}",
                "",
                f"- عدد الصفوف التقريبي: {inspection.max_row}",
                f"- عدد الأعمدة التقريبي: {inspection.max_column}",
                "",
                "### الأعمدة المكتشفة",
                "",
                "\n".join(f"- {header or f'Column {index + 1}'}" for index, header in enumerate(inspection.headers))
                or "- لا توجد أعمدة مكتشفة.",
                "",
                f"### أول {preview_rows} صفوف بعد صف الأعمدة",
                "",
                markdown_table(inspection.headers, inspection.preview_rows),
                "",
            ]
        )

    output_path.write_text("\n".join(sections), encoding="utf-8")
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Safely inspect an Excel file without modifying it.")
    parser.add_argument("--file", required=True, help="Path to .xlsx or .xlsm file to inspect.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Directory for the inspection report.")
    parser.add_argument("--preview-rows", type=int, default=5, help="Number of preview rows per sheet.")
    args = parser.parse_args()

    output_path = inspect_workbook(
        file_path=Path(args.file),
        output_dir=Path(args.output_dir),
        preview_rows=max(args.preview_rows, 0),
    )
    print(f"Excel inspection report generated: {output_path}")


if __name__ == "__main__":
    main()
