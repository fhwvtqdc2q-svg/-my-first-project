#!/usr/bin/env python3
"""
Convert the Arabic materials Excel file into a normalized products CSV.

This converter is intentionally conservative:
- It reads the Excel file in read-only mode.
- It does not modify the original Excel file.
- It writes a new CSV file under converted/ by default.
- It only converts fields that are clearly present in the inspected workbook.

Expected workbook shape from inspection:
- Sheet name: دليل المواد
- Column header: الاسم
- Column header: عامل تحويل الوحدة الثانية

Because the inspected file does not clearly contain stock quantities, prices,
or warehouse names, this script does not invent them. It leaves category, unit,
and price_usd blank and marks products as active.
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - user environment guard
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: python -m pip install -r requirements.txt"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_DIR = BASE_DIR / "converted"
DEFAULT_SHEET_NAME = "دليل المواد"


@dataclass
class ProductRow:
    product_code: str
    product_name: str
    category: str
    unit: str
    price_usd: str
    is_active: str
    secondary_unit_conversion_factor: str


def normalize_header(value: Any) -> str:
    return str(value or "").replace("\n", " ").replace("\r", " ").strip()


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def safe_product_code(index: int) -> str:
    return f"P{index:05d}"


def is_probably_group_header(name: str, conversion_factor: str) -> bool:
    """Skip group/title rows such as '1970' when they are not actual products."""
    if not name:
        return True
    if conversion_factor:
        return False
    # A row that is only digits is likely a group header based on the inspection sample.
    if re.fullmatch(r"\d+", name):
        return True
    return False


def get_header_map(header_values: Iterable[Any]) -> Dict[str, int]:
    headers = [normalize_header(value) for value in header_values]
    return {header: index for index, header in enumerate(headers) if header}


def require_column(header_map: Dict[str, int], column_name: str) -> int:
    if column_name not in header_map:
        available = ", ".join(header_map.keys()) or "none"
        raise ValueError(f"Required column '{column_name}' was not found. Available columns: {available}")
    return header_map[column_name]


def convert_materials_excel(
    file_path: Path,
    output_dir: Path,
    sheet_name: str,
    output_name: str,
) -> Path:
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Only .xlsx and .xlsm files are supported. Convert .xls to .xlsx first.")

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        workbook.close()
        raise ValueError(f"Sheet '{sheet_name}' was not found. Available sheets: {available}")

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        workbook.close()
        raise ValueError(f"Sheet '{sheet_name}' is empty.")

    header_map = get_header_map(header_row)
    name_index = require_column(header_map, "الاسم")
    conversion_index = header_map.get("عامل تحويل الوحدة الثانية")

    products: List[ProductRow] = []
    skipped_rows = 0
    seen_names: set[str] = set()

    for row in rows:
        name = cell_to_text(row[name_index] if name_index < len(row) else "")
        conversion_factor = ""
        if conversion_index is not None and conversion_index < len(row):
            conversion_factor = cell_to_text(row[conversion_index])

        if is_probably_group_header(name, conversion_factor):
            skipped_rows += 1
            continue

        # Avoid duplicate product names while preserving the first occurrence.
        normalized_name = re.sub(r"\s+", " ", name).strip()
        if normalized_name in seen_names:
            skipped_rows += 1
            continue
        seen_names.add(normalized_name)

        products.append(
            ProductRow(
                product_code=safe_product_code(len(products) + 1),
                product_name=normalized_name,
                category="",
                unit="",
                price_usd="",
                is_active="true",
                secondary_unit_conversion_factor=conversion_factor,
            )
        )

    workbook.close()

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / output_name
    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "product_code",
                "product_name",
                "category",
                "unit",
                "price_usd",
                "is_active",
                "secondary_unit_conversion_factor",
            ],
        )
        writer.writeheader()
        for product in products:
            writer.writerow(product.__dict__)

    summary_path = output_dir / "products_from_excel_summary.txt"
    summary_path.write_text(
        "\n".join(
            [
                "Materials Excel conversion summary",
                f"Source file: {file_path}",
                f"Sheet: {sheet_name}",
                f"Output CSV: {output_path}",
                f"Converted products: {len(products)}",
                f"Skipped rows: {skipped_rows}",
                "Original Excel file was not modified.",
            ]
        ),
        encoding="utf-8",
    )

    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert materials Excel workbook to normalized products CSV.")
    parser.add_argument("--file", required=True, help="Path to the materials .xlsx/.xlsm file.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Worksheet name to convert.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Directory for converted CSV files.")
    parser.add_argument("--output-name", default="products_from_excel.csv", help="Output CSV filename.")
    args = parser.parse_args()

    output_path = convert_materials_excel(
        file_path=Path(args.file),
        output_dir=Path(args.output_dir),
        sheet_name=args.sheet,
        output_name=args.output_name,
    )
    print(f"Products CSV generated: {output_path}")
    print("Original Excel file was not modified.")


if __name__ == "__main__":
    main()
