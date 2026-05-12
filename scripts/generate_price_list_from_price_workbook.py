#!/usr/bin/env python3
"""
Generate customer-facing USD and SYP price lists directly from the Arabic price workbook.

Expected workbook layout, based on the user's price model:
- Worksheet: نشرة اسعار $
- Repeated 3-column blocks:
  - اسم المادة
  - التوضيب
  - السعر $

Published list rules:
- Do not show group, code, quantity, or packaging/conversion factor.
- Show only: الصنف | الوحدة | السعر.
- If packaging/conversion factor is present and greater than 1:
  - USD list unit = كرتونة
  - SYP list unit = كروز
  - SYP price = USD carton price / factor * exchange rate
- If packaging/conversion factor is missing, zero, or one:
  - unit = قطعة
  - SYP price = USD piece price * exchange rate
- Items without USD price are written to a review section, not published in the price rows.
- The original Excel file is read-only and never modified.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with: python -m pip install -r requirements.txt"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent.parent
DEFAULT_OUTPUT_DIR = BASE_DIR / "output"
DEFAULT_SHEET_NAME = "نشرة اسعار $"


@dataclass
class PriceItem:
    group: str
    name: str
    factor: float
    price_usd: Optional[float]


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").replace("\r", " ").strip()


def parse_float(value: Any) -> Optional[float]:
    raw = text(value).replace(",", "")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def money(value: float) -> str:
    return f"{value:,.2f}"


def syp_money(value: float) -> str:
    return f"{round(value):,}"


def normalize_name(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def find_price_blocks(sheet: Any) -> List[Tuple[int, int, int, str, int]]:
    """Return tuples of (name_col, factor_col, price_col, group_name, header_row)."""
    blocks: List[Tuple[int, int, int, str, int]] = []
    max_scan_rows = min(sheet.max_row, 25)
    for row_idx in range(1, max_scan_rows + 1):
        for col_idx in range(1, sheet.max_column + 1):
            header = text(sheet.cell(row_idx, col_idx).value).replace(" ", "")
            if header != "اسمالمادة":
                continue
            factor_header = text(sheet.cell(row_idx, col_idx + 1).value)
            price_header = text(sheet.cell(row_idx, col_idx + 2).value)
            if "توضيب" not in factor_header or ("سعر" not in price_header and "$" not in price_header):
                continue

            group_name = "غير مصنف"
            for above in range(row_idx - 1, 0, -1):
                candidate = normalize_name(text(sheet.cell(above, col_idx).value))
                if candidate:
                    group_name = candidate
                    break
            blocks.append((col_idx, col_idx + 1, col_idx + 2, group_name, row_idx))
    return blocks


def load_items_from_workbook(file_path: Path, sheet_name: str) -> Tuple[List[PriceItem], List[str]]:
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Only .xlsx and .xlsm files are supported. Convert .xls to .xlsx first.")

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        workbook.close()
        raise ValueError(f"Sheet '{sheet_name}' was not found. Available sheets: {available}")

    sheet = workbook[sheet_name]
    blocks = find_price_blocks(sheet)
    if not blocks:
        workbook.close()
        raise ValueError("No price blocks were detected. Expected: اسم المادة / التوضيب / السعر $")

    items: List[PriceItem] = []
    review: List[str] = []
    seen: set[str] = set()

    for name_col, factor_col, price_col, group_name, header_row in blocks:
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            name = normalize_name(text(sheet.cell(row_idx, name_col).value))
            if not name:
                continue

            factor = parse_float(sheet.cell(row_idx, factor_col).value)
            if factor is None or factor <= 1:
                factor = 1.0

            price_usd = parse_float(sheet.cell(row_idx, price_col).value)
            dedupe_key = f"{group_name}|{name}"
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)

            if price_usd is None:
                review.append(f"{name}: لا يوجد له سعر بالدولار في نموذج الأسعار.")
                continue

            items.append(PriceItem(group=group_name, name=name, factor=factor, price_usd=price_usd))

    workbook.close()
    items.sort(key=lambda item: (item.group, item.name))
    review.sort()
    return items, review


def build_rows(items: List[PriceItem], exchange_rate: float) -> Tuple[str, str]:
    usd_rows: List[str] = []
    syp_rows: List[str] = []

    for item in items:
        if item.factor > 1:
            usd_unit = "كرتونة"
            syp_unit = "كروز"
            syp_price = (item.price_usd or 0) / item.factor * exchange_rate
        else:
            usd_unit = "قطعة"
            syp_unit = "قطعة"
            syp_price = (item.price_usd or 0) * exchange_rate

        usd_rows.append(f"| {item.name} | {usd_unit} | {money(item.price_usd or 0)} |")
        syp_rows.append(f"| {item.name} | {syp_unit} | {syp_money(syp_price)} |")

    if not usd_rows:
        usd_rows.append("| لا توجد أصناف قابلة للنشر | - | - |")
    if not syp_rows:
        syp_rows.append("| لا توجد أصناف قابلة للنشر | - | - |")
    return "\n".join(usd_rows), "\n".join(syp_rows)


def bullet_lines(items: List[str]) -> str:
    if not items:
        return "- لا توجد أصناف تحتاج تسعير."
    return "\n".join(f"- {item}" for item in items)


def generate_price_list_from_workbook(
    file_path: Path,
    output_dir: Path,
    sheet_name: str,
    report_date: date,
    exchange_rate: float,
) -> Path:
    if exchange_rate <= 0:
        raise ValueError("Exchange rate must be greater than zero.")

    items, review_items = load_items_from_workbook(file_path, sheet_name)
    usd_rows, syp_rows = build_rows(items, exchange_rate)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"price-list-from-workbook-{report_date.isoformat()}.md"
    content = "\n".join(
        [
            "# نشرة أسعار اليوم",
            "",
            f"التاريخ: {report_date.isoformat()}",
            f"سعر الصرف: {syp_money(exchange_rate)}",
            "",
            "## نشرة الدولار",
            "",
            "| الصنف | الوحدة | السعر بالدولار |",
            "|---|---|---:|",
            usd_rows,
            "",
            "## نشرة السوري",
            "",
            "| الصنف | الوحدة | السعر بالليرة السورية |",
            "|---|---|---:|",
            syp_rows,
            "",
            "## أصناف تحتاج تسعير قبل النشر",
            "",
            bullet_lines(review_items),
            "",
            "## ملاحظات داخلية",
            "",
            "- لا تظهر المجموعة أو الكود أو الكمية أو التوضيب في النشرة المنشورة.",
            "- التوضيب يستخدم داخليًا فقط لحساب سعر الكروز.",
            "- إذا لم يوجد توضيب، يعامل الصنف كقطعة.",
            "- يجب مراجعة الأسعار يدويًا قبل النشر النهائي.",
            f"- تم التوليد من الملف: `{file_path}`",
            f"- وقت التوليد: {datetime.now().isoformat(timespec='seconds')}",
            "- لم يتم تعديل ملف Excel الأصلي.",
            "",
        ]
    )
    output_path.write_text(content, encoding="utf-8")
    return output_path


def parse_date(value: str) -> date:
    return datetime.strptime(value.strip(), "%Y-%m-%d").date()


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate price list directly from the Arabic price workbook.")
    parser.add_argument("--file", required=True, help="Path to the .xlsx/.xlsm price workbook.")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Worksheet name to read.")
    parser.add_argument("--date", default=date.today().isoformat(), help="Report date in YYYY-MM-DD format.")
    parser.add_argument("--exchange-rate", required=True, type=float, help="USD to SYP exchange rate.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Directory for generated output.")
    args = parser.parse_args()

    output_path = generate_price_list_from_workbook(
        file_path=Path(args.file),
        output_dir=Path(args.output_dir),
        sheet_name=args.sheet,
        report_date=parse_date(args.date),
        exchange_rate=args.exchange_rate,
    )
    print(f"Price list generated from workbook: {output_path}")
    print("Original Excel file was not modified.")


if __name__ == "__main__":
    main()
